from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def read_report_1(path: Path, sheet_name: str = "Отчёт") -> Dict[str, Any]:

    if not path.exists():
        raise FileNotFoundError(f"report_1 not found: {path}")

    wb = load_workbook(path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet not found: {sheet_name}")

    ws = wb[sheet_name]

    return {
        "path": str(path),
        "workbook": wb,
        "ws": ws,
        "sheet_name": sheet_name,
        "sheetnames": wb.sheetnames,
    }


def read_report_2(path: Path) -> Dict[str, Any]:

    if not path.exists():
        raise FileNotFoundError(f"report_2 not found: {path}")

    wb = load_workbook(path, data_only=True)

    target_sheet = None

    for s in wb.sheetnames:

        low = s.strip().lower()

        if "распределение" in low and "нп" in low:
            target_sheet = s
            break

    if target_sheet is None:
        raise KeyError("sheet 'Распределение НП по РК' not found")

    ws = wb[target_sheet]

    return {
        "path": str(path),
        "workbook": wb,
        "ws": ws,
        "sheet_name": target_sheet,
        "sheetnames": wb.sheetnames,
    }


def write_output_test(out_path: Path, sheet_name: str = "Таблица1") -> None:

    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    ws = wb.active

    ws.title = sheet_name

    ws["A1"] = "OK"

    wb.save(out_path)


def _safe_sheet_title(name: str) -> str:

    bad = [":", "\\", "/", "?", "*", "[", "]"]

    for ch in bad:
        name = name.replace(ch, " ")

    name = name.strip()

    if len(name) > 31:
        name = name[:31]

    return name or "Sheet"


def _delete_default_sheet_if_needed(wb: Workbook) -> None:

    if len(wb.worksheets) == 1 and wb.worksheets[0].title == "Sheet":
        wb.remove(wb.worksheets[0])


def _set_col_widths(
    ws,
    headers: List[str],
    start_col: int = 1,
    max_rows_scan: int = 200,
) -> None:

    for i, h in enumerate(headers, start=start_col):

        letter = get_column_letter(i)

        max_len = len(str(h))

        for r in range(1, min(ws.max_row, max_rows_scan) + 1):

            v = ws.cell(row=r, column=i).value

            if v is None:
                continue

            max_len = max(max_len, len(str(v)))

        ws.column_dimensions[letter].width = min(
            max(10, max_len + 2),
            55,
        )


def _write_rows_as_table(
    ws,
    headers: List[str],
    rows: List[Dict[str, Any]],
    start_row: int = 1,
    start_col: int = 1,
) -> int:

    for j, h in enumerate(headers, start=start_col):

        c = ws.cell(row=start_row, column=j, value=h)

        c.font = Font(bold=True)

        c.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    r_out = start_row + 1

    for row in rows:

        for j, h in enumerate(headers, start=start_col):

            ws.cell(
                row=r_out,
                column=j,
                value=row.get(h, None),
            )

        r_out += 1

    return r_out


def _write_section_title(
    ws,
    title: str,
    row: int,
    start_col: int = 1,
    span_cols: int = 3,
) -> int:

    c = ws.cell(row=row, column=start_col, value=title)

    c.font = Font(bold=True)

    c.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    if span_cols > 1:

        ws.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=start_col + span_cols - 1,
        )

    return row + 1


def _write_table2(ws, t: Dict[str, Any]) -> None:

    export_rows = t.get("export_rows", []) or []
    trade_rows = t.get("trade_rows", []) or []

    headers = ["№", "Страна", "Количество перевозок"]

    row = 1

    row = _write_section_title(
        ws,
        "Статистические данные экспорта и взаимной торговли из РК",
        row,
        span_cols=3,
    )

    row = _write_section_title(
        ws,
        "Экспорт",
        row,
        span_cols=3,
    )

    row = _write_rows_as_table(
        ws,
        headers=headers,
        rows=export_rows,
        start_row=row,
    )

    row += 1

    row = _write_section_title(
        ws,
        "Взаимная торговля",
        row,
        span_cols=3,
    )

    row = _write_rows_as_table(
        ws,
        headers=headers,
        rows=trade_rows,
        start_row=row,
    )

    ws.freeze_panes = "A4"

    _set_col_widths(ws, headers)


def _write_table3(ws, t: Dict[str, Any]) -> None:

    auto_out_rows = t.get("auto_out_rows", []) or []
    auto_in_rows = t.get("auto_in_rows", []) or []

    rail_out_rows = t.get("rail_out_rows", []) or []
    rail_in_rows = t.get("rail_in_rows", []) or []

    headers = ["Направление", "Количество"]

    row = 1

    row = _write_section_title(
        ws,
        "Статистические данные завершенных перевозок",
        row,
        span_cols=len(headers),
    )

    row = _write_section_title(
        ws,
        "Завершённые автомобильные перевозки из Республики Казахстан",
        row,
        span_cols=len(headers),
    )

    row = _write_rows_as_table(
        ws,
        headers,
        auto_out_rows,
        row,
    )

    row += 1

    row = _write_section_title(
        ws,
        "Завершенные автомобильные перевозки в Республику Казахстан",
        row,
        span_cols=len(headers),
    )

    row = _write_rows_as_table(
        ws,
        headers,
        auto_in_rows,
        row,
    )

    row += 1

    row = _write_section_title(
        ws,
        "Завершённые железнодорожные перевозки из Республики Казахстан",
        row,
        span_cols=len(headers),
    )

    row = _write_rows_as_table(
        ws,
        headers,
        rail_out_rows,
        row,
    )

    row += 1

    row = _write_section_title(
        ws,
        "Завершенные железнодорожные перевозки в Республику Казахстан",
        row,
        span_cols=len(headers),
    )

    row = _write_rows_as_table(
        ws,
        headers,
        rail_in_rows,
        row,
    )

    ws.freeze_panes = "A4"

    _set_col_widths(ws, headers)


def _write_table4(ws, t: Dict[str, Any]) -> None:

    rows = t.get("rows", []) or []
    date_label = str(
        t.get("date_label_short") or t.get("date_label", "") or ""
    ).strip()

    headers = [
        "Пункты\nпропуска/региональ\nные склады",
        f"НП на {date_label}" if date_label else "НП",
    ]
    data_rows = [
        {
            headers[0]: row.get("Точка"),
            headers[1]: row.get("Количество НП"),
        }
        for row in rows
    ]

    row = _write_rows_as_table(
        ws,
        headers=headers,
        rows=data_rows,
        start_row=1,
    )

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_cells in ws.iter_rows(min_row=1, max_row=row - 1, min_col=1, max_col=2):
        for cell in row_cells:
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if cell.column == 2 else "left",
                vertical="center",
                wrap_text=True,
            )

            if cell.row == 1 or cell.column == 2:
                cell.font = Font(bold=True)

    ws.row_dimensions[1].height = 42
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    ws.freeze_panes = "A2"



def _write_table5(ws, t: Dict[str, Any]) -> None:

    rows = t.get("rows", []) or []

    headers = [
        "Страна начала перевозки",
        "Количество перевозок",
    ]

    row = 1

    row = _write_section_title(
        ws,
        "Статистика по перевозкам, направленным в сторону Республики Казахстан",
        row,
        span_cols=2,
    )

    _write_rows_as_table(
        ws,
        headers=headers,
        rows=rows,
        start_row=row,
    )

    ws.freeze_panes = "A3"

    _set_col_widths(ws, headers)


def write_output(out_path: Path, tables: List[Any]) -> None:

    out_path.parent.mkdir(parents=True, exist_ok=True)

    if not tables:

        write_output_test(
            out_path,
            sheet_name="Таблица1",
        )

        print(
            f"[io_excel.write_output] "
            f"tables_stub_count=0 "
            f"(created {out_path} with Таблица1:A1='OK')"
        )

        return

    wb = Workbook()

    _delete_default_sheet_if_needed(wb)

    for t in tables:

        if not isinstance(t, dict):
            continue

        sheet_title = _safe_sheet_title(
            str(t.get("table", "Sheet"))
        )

        ws = wb.create_sheet(title=sheet_title)

        if sheet_title == "Таблица2":

            _write_table2(ws, t)

            continue

        if sheet_title == "Таблица3":

            _write_table3(ws, t)

            continue

        if sheet_title == "Таблица4":

            _write_table4(ws, t)

            continue

        if sheet_title == "Таблица5":

            _write_table5(ws, t)

            continue

        rows = t.get("rows", [])

        if not isinstance(rows, list) or (
            rows and not isinstance(rows[0], dict)
        ):

            ws["A1"] = str(t)

            continue

        if sheet_title == "Таблица1":

            headers = [
                "ТП/ЖД",
                "РФ",
                "РБ",
                "КР",
                "РА",
                "Итого",
            ]

        else:

            headers = list(rows[0].keys()) if rows else ["data"]

        _write_rows_as_table(ws, headers, rows)

        ws.freeze_panes = "A2"

        _set_col_widths(ws, headers)

    wb.save(out_path)

    print(
        f"[io_excel.write_output] "
        f"wrote {len(tables)} table(s) to: {out_path}"
    )
